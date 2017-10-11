#-*- coding: utf-8 -*-
from django.shortcuts import render, redirect
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from uploads.core.models import Document
from uploads.core.forms import DocumentForm
from pyswip import *
from openpyxl import *
from uploads.core.Process import *



def home(request):
    documents = Document.objects.all()
    return render(request, 'core/home.html', { 'documents': documents })


def simple_upload(request):
    if request.method == 'POST' and request.FILES['myfile']:
        files=request.FILES.getlist('myfile')
        wb = Workbook()
        
        fs = FileSystemStorage()
        filename = fs.save(files[0].name, files[0])
        
        uploaded_file_url = fs.url(filename)
        wb2=load_workbook(files[0])
        
        wb=wb2.active
        name=wb["B5"].value
        idennum=wb["I5"].value
        day=wb["C9"].value
        contract=wb["B3"].value
        hospital=wb["C12"].value
        type1=wb["B10"].value

        wb3=load_workbook(files[1])
        wb=wb3.active
        name2=wb["B6"].value
        pay=wb["C22"].value


        nam=str(name.encode('utf-8'))
        nam2=str(name2.encode('utf-8'))
        con=str(contract.encode('utf-8'))   ##엑셀값을 가져오면 unicode로 utf-8로 바꿔줘야
        hos=str(hospital.encode('utf-8'))
        typ=str(type1.encode('utf-8'))
        if con=="우리가족상해보장공제":
            con="old"
        elif con=="무배당좋은이웃의료비보장공제":
            con="new"
        else:
            con="newnew"

        if hos[:2]=="의원":
            hos="small"
        elif hos[:4]=="종합병원":
            hos="middle"
        else :
            hos="big"

        if typ=="입원":
            typ="ipwon"
        elif typ=="통원":
            typ="tongwon"



        # Prolog
        pr=Insurance()
        pr.tell("Contract("+con+")")
        pr.tell("Type("+typ+")")
        value=pr.ask("Pay(a,b,c,d)")
        small=value.get(expr("a"))
        big=value.get(expr("b"))
        day=value.get(expr("c"))
        count=value.get(expr("d"))



        if pay>=small and pay<big:
            value="success"
            wb4 = Workbook()
            ws1 = wb4.active 
            ws1.title = "Report_Sheet" 
            ws1["A1"] = pay
            ws1["A2"] = name
            ws1["A3"] = con
            filename="Report.xlsx"
            wb4.save("media/"+filename) 
            uploaded_file_url = fs.url(filename)
            
            return render(request, 'core/simple_upload.html', {
            'uploaded_file_url':uploaded_file_url,
            'value':value,
        })


        if nam1!=nam2:
            return render(request, 'core/simple_upload.html')
        

            


        return render(request, 'core/simple_upload.html', {
            'uploaded_file_url':uploaded_file_url,
            'value':value,
        })

    return render(request, 'core/simple_upload.html')


def model_form_upload(request):
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        
        if form.is_valid():
            form.save()
            return redirect('home')
    else:
        form = DocumentForm()
    return render(request, 'core/model_form_upload.html', {
        'form': form
    })
